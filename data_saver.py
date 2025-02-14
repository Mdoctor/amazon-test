import os
import time
import re
import pandas as pd
from logger import logger
from openpyxl.styles import Font, PatternFill, Alignment


class DataSaver:
    # 定义输出目录
    OUTPUT_DIR = 'scraper_excel'

    @staticmethod
    def save_to_excel(products, category_name):
        """保存商品信息到Excel文件"""
        try:
            if not products:
                logger.warning("No products to save")
                return

            # 创建输出目录（如果不存在）
            os.makedirs(DataSaver.OUTPUT_DIR, exist_ok=True)

            # 准备数据
            safe_category_name = re.sub(r'[<>:"/\\|?*]', '_', category_name)
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            filename = f'{safe_category_name}_bestsellers_{timestamp}.xlsx'

            # 构建完整的文件路径
            full_path = os.path.join(DataSaver.OUTPUT_DIR, filename)

            # 创建DataFrame，包含所有字段
            df = pd.DataFrame([
                {
                    'Rank': i + 1,
                    'ASIN': p['asin'],
                    'Title': p['title'],
                    'Brand': p['brand'],
                    'Current Price': p['price']['current_price'],
                    'Original Price': p['price']['original_price'],
                    'Deal Price': p['price']['deal_price'],
                    'Price Range Min': p['price']['price_range']['min'],
                    'Price Range Max': p['price']['price_range']['max'],
                    'Savings Amount': p['price']['savings']['amount'],
                    'Savings Percentage': p['price']['savings']['percentage'],
                    'Prime Price': p['price']['prime_price'],
                    'Installment': p['price']['installment'],
                    'Coupon': p['price']['coupon'],
                    'Rating': p['rating'],
                    'Review Count': p['review_count'],
                    'Availability': p['availability'],
                    'Description': p['description'],
                    'Image URL': p['image_url'],
                    'Product URL': p['url'],
                    'Category': p['category'],
                    'Timestamp': p['timestamp']
                }
                for i, p in enumerate(products)
            ])

            # 直接保存到指定路径
            DataSaver._save_with_formatting(df, full_path)
            logger.info(f"Successfully saved to {full_path}")

        except Exception as e:
            logger.error(f"Error saving to Excel: {str(e)}")
            # 如果Excel保存失败，尝试CSV备份
            DataSaver._save_as_csv_backup(df, safe_category_name, timestamp)

    @staticmethod
    def _save_with_formatting(df, full_path):
        """保存Excel文件并设置格式"""
        try:
            # 在保存之前处理数据，限制长文本
            df_formatted = df.copy()

            # 创建 ExcelWriter 对象
            with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
                # 将处理后的数据写入Excel
                df_formatted.to_excel(writer, index=False, sheet_name='Bestsellers')
                worksheet = writer.sheets['Bestsellers']

                # 设置默认行高 (单位：点)
                default_row_height = 25  # 可以调整这个值来改变行高
                worksheet.sheet_format.defaultRowHeight = default_row_height

                # 设置所有行的高度一致
                for row_idx in range(1, worksheet.max_row + 1):
                    worksheet.row_dimensions[row_idx].height = default_row_height

                # 设置列宽
                for idx, col in enumerate(df_formatted.columns):

                    # 根据列类型设置宽度
                    if col in ['Description', 'Title', 'URL', 'Image URL', 'Product URL']:
                        adjusted_width = 40
                    elif any(price_text in col for price_text in ['Price', 'Savings']):
                        adjusted_width = 20
                    elif col in ['Rank', 'Review Count']:
                        adjusted_width = 20
                    else:
                        adjusted_width = 15

                    # 设置列宽
                    column_letter = chr(65 + idx) if idx < 26 else chr(64 + idx // 26) + chr(65 + (idx % 26))
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # 设置标题行格式
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # 设置数据行格式
                data_alignment = Alignment(vertical='center', wrap_text=False)
                price_alignment = Alignment(horizontal='right', vertical='center', wrap_text=False)

                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        # 设置基本对齐方式
                        cell.alignment = data_alignment

                        # 特殊列的处理
                        column_name = df_formatted.columns[cell.column - 1]

                        # 价格相关列右对齐
                        if any(price_text in column_name for price_text in ['Price', 'Savings']):
                            cell.alignment = price_alignment

                        # URL列添加超链接
                        if column_name in ['Product URL', 'Image URL']:
                            if cell.value and cell.value != 'N/A' and '...' not in str(cell.value):
                                cell.hyperlink = cell.value
                                cell.font = Font(color="0000FF", underline="single")

                        # 数字列居中对齐
                        if column_name in ['Rank', 'Review Count']:
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

                # 设置自动筛选
                worksheet.auto_filter.ref = worksheet.dimensions

                # 冻结标题行
                worksheet.freeze_panes = 'A2'

        except Exception as e:
            logger.error(f"Error formatting Excel file: {str(e)}")
            # 如果格式化失败，尝试基本保存
            df.to_excel(full_path, index=False)

    @staticmethod
    def _save_as_csv_backup(df, category_name, timestamp):
        """作为备份保存为CSV文件"""
        try:
            # 确保输出目录存在
            os.makedirs(DataSaver.OUTPUT_DIR, exist_ok=True)

            # 构建CSV文件路径
            csv_filename = f'{category_name}_bestsellers_{timestamp}.csv'
            csv_path = os.path.join(DataSaver.OUTPUT_DIR, csv_filename)

            # 保存CSV文件
            df.to_csv(csv_path, index=False, encoding='utf-8-sig')
            logger.info(f"Saved data as CSV backup at: {csv_path}")
        except Exception as csv_error:
            logger.error(f"Failed to save as CSV as well: {str(csv_error)}")